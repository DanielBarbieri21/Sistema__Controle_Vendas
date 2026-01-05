import pandas as pd
import sqlite3
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def exportar_excel(dados, caminho_arquivo, incluir_formatacao=True):
    """
    Exporta dados para Excel com formatação profissional
    
    Args:
        dados: Lista de tuplas com os dados (id, cliente, produto, quantidade, valor, data, numero_venda)
        caminho_arquivo: Caminho onde salvar o arquivo Excel
        incluir_formatacao: Se True, aplica formatação profissional ao arquivo
    """
    if not dados:
        raise ValueError("Não há dados para exportar")
    
    # Criar DataFrame
    df = pd.DataFrame(dados, columns=[
        'ID', 'Cliente', 'Produto', 'Quantidade', 'Valor Total', 'Data Venda', 'Número Venda'
    ])
    
    # Remover coluna ID (não precisa no Excel)
    df = df.drop('ID', axis=1)
    
    # Formatar valores monetários
    df['Valor Total'] = df['Valor Total'].apply(lambda x: f"R$ {float(x):.2f}" if x else "R$ 0.00")
    
    # Salvar para Excel
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Vendas', index=False)
        
        if incluir_formatacao:
            # Obter a planilha
            workbook = writer.book
            worksheet = writer.sheets['Vendas']
            
            # Estilos
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Formatar cabeçalho
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style
            
            # Ajustar largura das colunas
            column_widths = {
                'A': 30,  # Cliente
                'B': 50,  # Produto
                'C': 12,  # Quantidade
                'D': 15,  # Valor Total
                'E': 15,  # Data Venda
                'F': 20   # Número Venda
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatar células de dados
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border_style
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Alinhar valores numéricos
                    if cell.column == 3:  # Quantidade
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif cell.column == 4:  # Valor Total
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif cell.column == 5:  # Data Venda
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adicionar linha de total
            total_row = worksheet.max_row + 2
            worksheet.cell(row=total_row, column=1, value="TOTAL GERAL").font = Font(bold=True, size=11)
            worksheet.cell(row=total_row, column=1).fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )
            
            # Calcular total
            total_valor = sum(float(str(cell.value).replace('R$', '').replace(' ', '').replace(',', '.')) 
                            for cell in worksheet['D'][1:worksheet.max_row] 
                            if cell.value and 'R$' in str(cell.value))
            
            worksheet.cell(row=total_row, column=4, value=f"R$ {total_valor:.2f}")
            worksheet.cell(row=total_row, column=4).font = Font(bold=True, size=11)
            worksheet.cell(row=total_row, column=4).fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )
            worksheet.cell(row=total_row, column=4).alignment = Alignment(horizontal='right', vertical='center')
            
            # Congelar primeira linha
            worksheet.freeze_panes = 'A2'
            
            # Adicionar informações no rodapé
            info_row = worksheet.max_row + 2
            worksheet.cell(row=info_row, column=1, 
                         value=f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            worksheet.cell(row=info_row, column=1).font = Font(italic=True, size=9, color="666666")
            
            # Aplicar bordas na linha de total
            for col in range(1, 7):
                cell = worksheet.cell(row=total_row, column=col)
                cell.border = border_style

def exportar_excel_completo(caminho_arquivo, filtro_data_inicio=None, filtro_data_fim=None, 
                            filtro_cliente=None, filtro_produto=None, incluir_formatacao=True):
    """
    Exporta dados completos do banco para Excel aplicando filtros
    
    Args:
        caminho_arquivo: Caminho onde salvar o arquivo Excel
        filtro_data_inicio: Data inicial (QDate ou None)
        filtro_data_fim: Data final (QDate ou None)
        filtro_cliente: Nome do cliente para filtrar (str ou None)
        filtro_produto: Nome do produto para filtrar (str ou None)
        incluir_formatacao: Se True, aplica formatação profissional
    """
    conn = sqlite3.connect("vendas.db")
    cursor = conn.cursor()
    
    query = """
        SELECT v.id, v.cliente_nome, i.produto, i.quantidade, 
               v.valor_total, v.data_venda, v.numero_venda,
               v.cliente_cidade, v.cliente_estado, v.status_venda, i.preco_unitario
        FROM vendas v
        JOIN itens_venda i ON i.venda_id = v.id
        WHERE 1=1
    """
    
    params = []
    
    if filtro_data_inicio:
        query += " AND (v.data_venda IS NULL OR v.data_venda >= ?)"
        params.append(filtro_data_inicio.toString("yyyy-MM-dd"))
    
    if filtro_data_fim:
        query += " AND (v.data_venda IS NULL OR v.data_venda <= ?)"
        params.append(filtro_data_fim.toString("yyyy-MM-dd"))
    
    if filtro_cliente:
        query += " AND v.cliente_nome LIKE ?"
        params.append(f"%{filtro_cliente}%")
    
    if filtro_produto:
        query += " AND i.produto LIKE ?"
        params.append(f"%{filtro_produto}%")
    
    cursor.execute(query, params)
    dados = cursor.fetchall()
    conn.close()
    
    if not dados:
        raise ValueError("Não há dados para exportar com os filtros aplicados")
    
    # Criar múltiplas abas no Excel
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        # Aba 1: Dados completos
        df_completo = pd.DataFrame(dados, columns=[
            'ID', 'Cliente', 'Produto', 'Quantidade', 'Valor Total', 'Data Venda', 
            'Número Venda', 'Cidade', 'Estado', 'Status', 'Preço Unitário'
        ])
        df_completo = df_completo.drop('ID', axis=1)
        df_completo['Valor Total'] = df_completo['Valor Total'].apply(
            lambda x: f"R$ {float(x):.2f}" if x else "R$ 0.00"
        )
        df_completo['Preço Unitário'] = df_completo['Preço Unitário'].apply(
            lambda x: f"R$ {float(x):.2f}" if x else "R$ 0.00"
        )
        df_completo.to_excel(writer, sheet_name='Vendas Detalhadas', index=False)
        
        # Aba 2: Resumo por cliente
        # Criar DataFrame temporário com valores numéricos para ordenação
        df_temp_cliente = df_completo.copy()
        df_temp_cliente['Valor Num'] = df_temp_cliente['Valor Total'].apply(
            lambda x: float(str(x).replace('R$', '').replace(' ', '').replace(',', '.')) if isinstance(x, str) else float(x) if x else 0
        )
        
        df_resumo_cliente = df_temp_cliente.groupby('Cliente').agg({
            'Valor Num': 'sum',
            'Quantidade': 'sum'
        }).reset_index()
        df_resumo_cliente.columns = ['Cliente', 'Total Vendido', 'Quantidade Total']
        df_resumo_cliente = df_resumo_cliente.sort_values('Total Vendido', ascending=False)
        df_resumo_cliente['Total Vendido'] = df_resumo_cliente['Total Vendido'].apply(
            lambda x: f"R$ {x:.2f}"
        )
        df_resumo_cliente.to_excel(writer, sheet_name='Resumo por Cliente', index=False)
        
        # Aba 3: Resumo por produto
        df_temp_produto = df_completo.copy()
        df_temp_produto['Valor Num'] = df_temp_produto['Valor Total'].apply(
            lambda x: float(str(x).replace('R$', '').replace(' ', '').replace(',', '.')) if isinstance(x, str) else float(x) if x else 0
        )
        
        df_resumo_produto = df_temp_produto.groupby('Produto').agg({
            'Valor Num': 'sum',
            'Quantidade': 'sum'
        }).reset_index()
        df_resumo_produto.columns = ['Produto', 'Total Vendido', 'Quantidade Total']
        df_resumo_produto = df_resumo_produto.sort_values('Total Vendido', ascending=False)
        df_resumo_produto['Total Vendido'] = df_resumo_produto['Total Vendido'].apply(
            lambda x: f"R$ {x:.2f}"
        )
        df_resumo_produto.to_excel(writer, sheet_name='Resumo por Produto', index=False)
        
        if incluir_formatacao:
            # Formatar todas as abas
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Formatar cabeçalho
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_style
                
                # Ajustar largura das colunas automaticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatar células de dados
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = border_style
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                worksheet.freeze_panes = 'A2'

